# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Тег, Арабский термин (из morphology-terms-ar.json), Английский, Русский (черновик), Спорный?
rows = [
    # --- Именные / знаменательные части речи ---
    ("N",         "اسم",                       "noun",                        "имя существительное", "НЕТ"),
    ("PN",        "علم",                       "proper noun",                 "имя собственное", "НЕТ"),
    ("ADJ",       "نعت",                       "adjective",                   "прилагательное (определение, на‘т)", "НЕТ"),
    ("V",         "فعل",                       "verb",                        "глагол", "НЕТ"),
    ("PRON",      "ضمير",                      "personal pronoun",            "личное местоимение", "НЕТ"),
    ("DEM",       "اسم اشارة",                 "demonstrative pronoun",       "указательное местоимение", "НЕТ"),
    ("REL",       "اسم موصول",                 "relative pronoun",            "относительное местоимение (союзное имя)", "ДА"),
    ("T",         "ظرف زمان",                  "time adverb",                 "обстоятельство времени (зарф заман)", "НЕТ"),
    ("LOC",       "ظرف مكان",                  "location adverb",             "обстоятельство места (зарф макан)", "НЕТ"),
    ("ACT_PCPL",  "اسم فاعل",                  "active participle",           "действительное причастие (имя деятеля, исм фа‘иль)", "ДА"),
    ("PASS_PCPL", "اسم مفعول",                 "passive participle",          "страдательное причастие (исм маф‘уль)", "ДА"),
    ("VN",        "مصدر",                      "verbal noun",                 "отглагольное имя / масдар", "ДА"),
    ("NV",        "اسم فعل",                   "verbal noun (ism fi‘l)",      "имя глагола (исм фи‘ль, напр. «هَيْهَاتَ»)", "ДА"),
    # --- Служебные частицы / харфы ---
    ("P",         "حرف جر",                    "preposition",                 "предлог (частица род. падежа, харф джарр)", "НЕТ"),
    ("DET",       "ال",                        "determiner",                  "определённый артикль («аль-»)", "НЕТ"),
    ("CONJ",      "حرف عطف",                   "coordinating conjunction",    "соединительный союз (харф ‘атф)", "НЕТ"),
    ("NEG",       "حرف نفي",                   "negative particle",           "отрицательная частица (харф нафй)", "НЕТ"),
    ("INTG",      "استفهامية",                 "interrogative particle",      "вопросительная частица (истифхам)", "НЕТ"),
    ("COND",      "شرطية",                     "conditional particle",        "условная частица (харф шарт)", "НЕТ"),
    ("VOC",       "حرف نداء",                  "vocative particle",           "частица обращения / звательная («о!», нида)", "НЕТ"),
    ("PRO",       "حرف نهي",                   "prohibition particle",        "частица запрета («ля» запрещающее, нахй)", "НЕТ"),
    ("FUT",       "حرف استقبال",               "future particle",             "частица будущего времени («са-/сауфа»)", "НЕТ"),
    ("EMPH",      "لام التوكيد",               "emphatic lām",                "усилительный лям (лям ат-таукид)", "НЕТ"),
    ("ACC",       "حرف نصب",                   "accusative particle",         "частица насба (винит. наклонения глагола)", "ДА"),
    ("INL",       "حروف مقطعة",                "Quranic initials",            "разрозненные буквы (хуруф мукатта‘а, в начале сур)", "ДА"),
    ("SUB",       "حرف مصدري",                 "subordinating conjunction",   "масдарная (изъяснительная) частица («ан/анна»)", "ДА"),
    ("REM",       "حرف استئنافية",             "resumption particle",         "частица возобновления / начала нового предложения (исти’наф)", "ДА"),
    ("RES",       "أداة حصر",                  "restriction particle",        "частица ограничения / исключительности («только, лишь», хаср)", "ДА"),
    ("RSLT",      "حرف واقع في جواب الشرط",    "result particle",             "частица ответа на условие (джаваб аш-шарт, «то…»)", "ДА"),
    ("AMD",       "حرف استدراك",               "amendment particle",          "частица оговорки / восполнения («однако, но», истидрак)", "ДА"),
    ("CIRC",      "حرف حال",                   "circumstantial particle",     "частица обстоятельства состояния («в то время как», халь)", "ДА"),
    ("EXP",       "أداة استثناء",              "exceptive particle",          "частица исключения («кроме, помимо», истисна)", "ДА"),
    ("INC",       "حرف ابتداء",                "inceptive particle",          "начинательная частица (ибтида, «ведь, вот»)", "ДА"),
    ("CAUS",      "حرف سببية",                 "particle of cause",           "частица причины / следствия («чтобы, так что», сабабийя)", "ДА"),
    ("CERT",      "حرف تحقيق",                 "particle of certainty",       "частица подтверждения («поистине, уже» — «кад»)", "ДА"),
    ("PREV",      "حرف كاف",                   "preventive particle",         "ограничительная частица «каффа» (напр. «иннама»)", "ДА"),
    ("RET",       "حرف اضراب",                 "retraction particle",         "частица перехода / отмены («но нет; напротив» — «баль»)", "ДА"),
    ("SUR",       "حرف فجاءة",                 "surprise particle",           "частица внезапности («и вот; как вдруг» — «иза» фуджа’ийя)", "ДА"),
    ("AVR",       "حرف ردع",                   "aversion particle",           "частица отрезвления / упрёка («так нет же!» — «калля», рад‘)", "ДА"),
    ("EQ",        "حرف تسوية",                 "equalization particle",       "частица уравнивания («всё равно… или», тасвийя)", "ДА"),
    ("COM",       "واو المعية",                "comitative particle",         "вав совместности (вав аль-ма‘ийя, «вместе с»)", "ДА"),
    ("EXL",       "حرف تفصيل",                 "explanation particle",        "частица разделения / детализации (тафсиль)", "ДА"),
    ("ANS",       "حرف جواب",                  "answer particle",             "частица ответа («да; конечно» — «на‘ам, балā», джаваб)", "ДА"),
    ("EXH",       "حرف تحضيض",                 "exhortation particle",        "частица побуждения / увещевания («почему бы не…», тахдид)", "ДА"),
    ("INT",       "حرف تفسير",                 "interpretation particle",     "частица пояснения («то есть, а именно», тафсир)", "ДА"),
    ("PRP",       "لام التعليل",               "purpose lām",                 "лям цели / причины («чтобы», лям ат-та‘лиль)", "ДА"),
    ("SUP",       "حرف زائد",                  "supplemental particle",       "избыточная (плеонастическая) частица (харф за’ид)", "ДА"),
    ("IMPV",      "لام الامر",                 "imperative lām",              "повелительный лям (лям аль-амр; ср. омонимию с повелит. наклонением)", "ДА"),
    ("ATT",       "حرف تنبيه",                 "particle of attention",       "частица привлечения внимания («о вот!; внемли» — «аля/ха», танбих)", "ДА"),
    ("DIST",      "لام البعد",                 "distance lām",                "лям дальности (в указат. «тот» — «зāлика»)", "ДА"),
    ("ADDR",      "حرف خطاب",                  "particle of address",         "каф обращения (харф хитаб, в указат. местоимениях — «-ка»)", "ДА"),
]

wb = Workbook()
ws = wb.active
ws.title = "Термины"
headers = ["Тег", "Арабский термин", "Английский", "Русский (черновик)", "Спорный?"]
ws.append(headers)

hdr_fill = PatternFill("solid", fgColor="305496")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
yes_fill = PatternFill("solid", fgColor="FFF2CC")  # подсветка спорных

for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for r in rows:
    ws.append(r)

for ri in range(2, ws.max_row + 1):
    disputed = ws.cell(ri, 5).value == "ДА"
    for ci in range(1, 6):
        cell = ws.cell(ri, ci)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(ci == 4))
        if ci == 1:
            cell.font = Font(bold=True, name="Consolas")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if ci == 2:
            cell.font = Font(size=14, name="Arial")
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if ci == 5:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="C00000" if disputed else "548235")
        if disputed:
            cell.fill = yes_fill

widths = [10, 24, 26, 60, 11]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{ws.max_row}"

out = "/home/user/tafsir-app/terms.xlsx"
wb.save(out)
print("rows:", len(rows))
print("disputed:", sum(1 for r in rows if r[4] == "ДА"))
print("saved:", out)
