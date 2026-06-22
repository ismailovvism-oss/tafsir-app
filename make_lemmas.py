# -*- coding: utf-8 -*-
import sys, json, unicodedata, re
sys.path.insert(0, "/tmp"); import ru
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open("/tmp/lemmas.json"))
def nk(s): return unicodedata.normalize("NFKC", s)
look = {}
for k, v in ru.RU.items():
    look.setdefault(nk(k), v)
def rus(l):
    return ru.RU.get(l) or look.get(nk(l), "")

# --- классификатор ---
FUNC_POS = {"P","ACC","INTG","REL","CONJ","COND","NEG","ANS","SUB","DEM","EXL",
            "RES","PRON","CERT","AMD","EXH","AVR","RET","ATT","FUT","T","LOC",
            "INL","EMPH","CAUS","CIRC","EQ","EXP","INC","INT","PREV","PRP","REM",
            "RSLT","SUP","SUR","VOC","PRO","COM"}

THEO_ROOTS = set("""أله ربب رحم أمن كفر شرك نفق سلم صلو زكو صوم حجج عبد شهد غيب وقي
هدي ضلل غفر توب رسل نبأ وحي حقق عذب أجر ثوب خلد بعث عرش سبح حمد ذكر دعو صبر فسق
ظلم طهر حرم حلل رزق برك قدس نور روح ملك كتب أيي جنن سجد ركع خلق صدق كذب قسط فلح
نجو فدي شفع قدر حكم عزز جبر قهر رأف ودد حفظ لطف كرم خشع قنت رهب نعم جزي حشر صرط
رضو سخط لعن برر جهد وعد ذنب صفح عفو حيي موت دين""".split())

THEO_LEMMAS = set(nk(x) for x in """قِيامَة عالَم عَلِيم عِلْم عَلّام عالِم أَعْلَم دُنْيا آخِر
شَيْطان قُرْءان فُرْقان مَلَك إِبْلِيس تَوْراة إِنجِيل زَبُور إِسْلام جِبْرِيل مِيكال مارُوت
هارُوت طاغُوت جاهِلِيَّة مَجُوس صابِئ يَهُود هادُ مَسِيح حَنِيف مُتَّقي إِيمان كَلِمَة
رَبّانِيِّن رِبِّيّ قُدُّوس قُدُس""".split())

USUS_LEMMAS = set(nk(x) for x in "صَلاة زَكاة صِيام صَوْم حَجّ حِجّ عُمْرَة رِبا رِب صَدَقَة".split())

STOP = set("""a an the of to and or in on for is are be by with not from at as into he she
it they you we i him her them us me his their your our my its who whom which that this these
those o and not (of) (the) be""".split())

def is_phrase(en):
    if not en: return False
    g = en[0].lower()
    g = re.sub(r"\([^)]*\)", " ", g)        # убрать скобочные пояснения
    g = re.sub(r"[^a-z\s'-]", " ", g)
    toks = [t for t in g.split() if t and t not in STOP]
    return len(set(toks)) >= 2

def classify(l, e):
    root = e["root"]; pos = e["pos"]; freq = e["freq"]; en = e["en"]
    nl = nk(l)
    if nl in USUS_LEMMAS:
        return "узус", "ДА"
    if root in THEO_ROOTS or nl in THEO_LEMMAS:
        return "теолог.", "ДА"
    if pos in FUNC_POS:
        return "частица", "ДА"
    if freq <= 2:
        return "редкое", "ДА"
    if is_phrase(en):
        return "рассинхрон", "ДА"
    return "обычное", "НЕТ"

rows = []
for l, e in d.items():
    cat, na = classify(l, e)
    rows.append({
        "lemma": l,
        "root": e["root"],
        "en": " / ".join(e["en"]),
        "ru": rus(l),
        "cat": cat,
        "na": na,
        "freq": e["freq"],
        "ex": ", ".join(e["ex"]),
    })

# сортировка: ДА сверху, далее по частоте (убыв.), затем по категории
rows.sort(key=lambda r: (r["na"] == "НЕТ", -r["freq"], r["cat"]))

wb = Workbook(); ws = wb.active; ws.title = "Леммы"
headers = ["Лемма (араб.)","Корень","Англ. глосс","Русский (черновик)",
           "Категория","На проверку?","Частота","Примеры (адреса)"]
ws.append(headers)

hdr_fill = PatternFill("solid", fgColor="305496")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D9D9D9"); border = Border(thin,thin,thin,thin)
cat_fill = {
    "теолог.":  PatternFill("solid", fgColor="FCE4D6"),
    "узус":     PatternFill("solid", fgColor="FFF2CC"),
    "частица":  PatternFill("solid", fgColor="DDEBF7"),
    "редкое":   PatternFill("solid", fgColor="EDEDED"),
    "рассинхрон":PatternFill("solid", fgColor="E2EFDA"),
    "обычное":  PatternFill("solid", fgColor="FFFFFF"),
}
for c,h in enumerate(headers,1):
    cell=ws.cell(1,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=border
    cell.alignment=Alignment(horizontal="center",vertical="center")

for r in rows:
    ws.append([r["lemma"],r["root"],r["en"],r["ru"],r["cat"],r["na"],r["freq"],r["ex"]])

for ri in range(2, ws.max_row+1):
    cat = ws.cell(ri,5).value
    da = ws.cell(ri,6).value == "ДА"
    for ci in range(1,9):
        cell=ws.cell(ri,ci); cell.border=border
        cell.alignment=Alignment(vertical="center", wrap_text=(ci in (3,4)))
        if ci in (1,2): cell.font=Font(size=14, name="Arial"); cell.alignment=Alignment(horizontal="right",vertical="center")
        if ci==6:
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.font=Font(bold=True, color="C00000" if da else "548235")
        if ci==7: cell.alignment=Alignment(horizontal="center",vertical="center")
        if ci==5: cell.fill=cat_fill.get(cat, cat_fill["обычное"])

widths=[16,12,34,46,13,13,9,18]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:H{ws.max_row}"

out="/home/user/tafsir-app/lemmas.xlsx"; wb.save(out)

from collections import Counter
cc=Counter(r["cat"] for r in rows); na=Counter(r["na"] for r in rows)
print("rows:",len(rows))
print("на проверку:",dict(na))
print("категории:",dict(cc))
print("saved:",out)
